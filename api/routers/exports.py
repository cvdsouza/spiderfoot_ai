"""Export API routes for CSV, Excel, JSON, and GEXF downloads."""

import csv
import json
import logging
import string
import time
from io import BytesIO, StringIO

import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response

from api.dependencies import get_config, get_db
from api.middleware.auth import require_permission
from spiderfoot import SpiderFootDb, SpiderFootHelpers

log = logging.getLogger(f"spiderfoot.{__name__}")

router = APIRouter(tags=["exports"])


def build_excel(data: list, column_names: list, sheet_name_index: int = 0) -> bytes:
    """Build an Excel workbook from data.

    Args:
        data: list of rows
        column_names: column header names
        sheet_name_index: which column to use as sheet name

    Returns:
        Excel file as bytes
    """
    row_nums = {}
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    column_names.pop(sheet_name_index)
    allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'

    for row in data:
        sheet_name = "".join([c for c in str(row.pop(sheet_name_index)) if c.upper() in allowed_sheet_chars])
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            for col_num, column_title in enumerate(column_names, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = column_title
            row_nums[sheet_name] = 2

        for col_num, cell_value in enumerate(row, 1):
            cell = sheet.cell(row=row_nums[sheet_name], column=col_num)
            cell.value = cell_value
        row_nums[sheet_name] += 1

    if row_nums:
        workbook.remove(default_sheet)

    workbook._sheets.sort(key=lambda ws: ws.title)

    with BytesIO() as f:
        workbook.save(f)
        f.seek(0)
        return f.read()


@router.get("/scans/{scan_id}/export/logs")
def export_scan_logs(scan_id: str, dialect: str = "excel", user: dict = Depends(require_permission("results", "read")), dbh: SpiderFootDb = Depends(get_db)):
    """Export scan logs as CSV."""
    try:
        data = dbh.scanLogs(scan_id, None, None, True)
    except Exception:
        raise HTTPException(status_code=404, detail="Scan not found")

    if not data:
        raise HTTPException(status_code=404, detail="Scan not found")

    fileobj = StringIO()
    parser = csv.writer(fileobj, dialect=dialect)
    parser.writerow(["Date", "Component", "Type", "Event", "Event ID"])
    for row in data:
        parser.writerow([
            time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000)),
            str(row[1]),
            str(row[2]),
            str(row[3]),
            row[4]
        ])

    return Response(
        content=fileobj.getvalue().encode('utf-8'),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}.log.csv",
            "Pragma": "no-cache",
        },
    )


@router.get("/scans/{scan_id}/export/correlations")
def export_correlations(
    scan_id: str,
    filetype: str = "csv",
    dialect: str = "excel",
    user: dict = Depends(require_permission("results", "read")),
    dbh: SpiderFootDb = Depends(get_db),
):
    """Export scan correlations as CSV or Excel."""
    try:
        data = dbh.scanCorrelationList(scan_id)
    except Exception:
        raise HTTPException(status_code=404, detail="Scan not found")

    if not data:
        raise HTTPException(status_code=404, detail="No correlations found")

    column_names = ["Title", "Risk", "Rule", "Events"]
    rows = []
    for row in data:
        rows.append([row[1], row[3], row[2], row[7]])

    if filetype in ("xlsx", "excel"):
        excel_data = build_excel(rows, column_names[:])
        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}-correlations.xlsx"},
        )

    fileobj = StringIO()
    parser = csv.writer(fileobj, dialect=dialect)
    parser.writerow(column_names)
    for row in rows:
        parser.writerow(row)

    return Response(
        content=fileobj.getvalue().encode('utf-8'),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}-correlations.csv",
            "Pragma": "no-cache",
        },
    )


@router.get("/scans/{scan_id}/export/events")
def export_events(
    scan_id: str,
    type: str = "ALL",
    filetype: str = "csv",
    dialect: str = "excel",
    user: dict = Depends(require_permission("results", "read")),
    dbh: SpiderFootDb = Depends(get_db),
):
    """Export scan events as CSV or Excel."""
    try:
        data = dbh.scanResultEvent(scan_id, type, filterFp=True)
    except Exception:
        raise HTTPException(status_code=404, detail="Scan not found")

    if not data:
        raise HTTPException(status_code=404, detail="No events found")

    column_names = ["Updated", "Type", "Module", "Source", "F/P", "Data"]
    rows = []
    for row in data:
        lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
        rows.append([lastseen, row[4], row[3], row[2], row[13], row[1]])

    if filetype in ("xlsx", "excel"):
        excel_data = build_excel(rows, column_names[:], sheet_name_index=1)
        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}.xlsx"},
        )

    fileobj = StringIO()
    parser = csv.writer(fileobj, dialect=dialect)
    parser.writerow(column_names)
    for row in rows:
        parser.writerow(row)

    return Response(
        content=fileobj.getvalue().encode('utf-8'),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}.csv",
            "Pragma": "no-cache",
        },
    )


@router.get("/scans/export/json")
def export_json_multi(ids: str, user: dict = Depends(require_permission("results", "read")), dbh: SpiderFootDb = Depends(get_db)):
    """Export multiple scans as JSON."""
    scan_ids = ids.split(',')
    all_data = {}

    for scan_id in scan_ids:
        try:
            data = dbh.scanResultEvent(scan_id, filterFp=True)
            res = dbh.scanInstanceGet(scan_id)
        except Exception:
            continue

        if not res:
            continue

        scan_data = []
        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            scan_data.append({
                "updated": lastseen,
                "type": row[4],
                "module": row[3],
                "source": row[2],
                "data": row[1],
            })

        all_data[scan_id] = {
            "name": res[0],
            "target": res[1],
            "results": scan_data,
        }

    return Response(
        content=json.dumps(all_data, indent=2).encode('utf-8'),
        media_type="application/json",
        headers={
            "Content-Disposition": "attachment; filename=SpiderFoot-export.json",
            "Pragma": "no-cache",
        },
    )


@router.get("/scans/export/graph")
def export_graph_multi(ids: str, user: dict = Depends(require_permission("results", "read")), dbh: SpiderFootDb = Depends(get_db)):
    """Export multiple scans as GEXF graph."""
    scan_ids = ids.split(',')
    all_data = []
    all_roots = []

    for scan_id in scan_ids:
        try:
            data = dbh.scanResultEvent(scan_id, filterFp=True)
            res = dbh.scanInstanceGet(scan_id)
        except Exception:
            continue

        if not res:
            continue

        all_roots.append(res[1])
        all_data.extend(data)

    gexf = SpiderFootHelpers.buildGraphGexf(all_roots, "SpiderFoot Export", all_data)

    return Response(
        content=gexf,
        media_type="application/xml",
        headers={
            "Content-Disposition": "attachment; filename=SpiderFoot-export.gexf",
            "Pragma": "no-cache",
        },
    )


@router.get("/search/export")
def export_search_results(
    id: str = "",
    eventType: str = "",
    value: str = "",
    filetype: str = "csv",
    dialect: str = "excel",
    user: dict = Depends(require_permission("results", "read")),
    dbh: SpiderFootDb = Depends(get_db),
):
    """Export search results as CSV or Excel."""
    if not value:
        value = ''

    regex = ""
    if value.startswith("/") and value.endswith("/"):
        regex = value[1:len(value) - 1]
        value = ""

    value = value.replace('*', '%')
    if value in [None, ""] and regex in [None, ""]:
        value = "%"
        regex = ""

    criteria = {
        'scan_id': id,
        'type': eventType,
        'value': value,
        'regex': regex,
    }

    try:
        data = dbh.search(criteria)
    except Exception:
        raise HTTPException(status_code=500, detail="Search failed")

    column_names = ["Updated", "Type", "Module", "Source", "Data"]
    rows = []
    for row in data:
        lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
        rows.append([lastseen, row[4], row[3], row[2], row[1]])

    if filetype in ("xlsx", "excel"):
        excel_data = build_excel(rows, column_names[:], sheet_name_index=1)
        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=SpiderFoot-search.xlsx"},
        )

    fileobj = StringIO()
    parser = csv.writer(fileobj, dialect=dialect)
    parser.writerow(column_names)
    for row in rows:
        parser.writerow(row)

    return Response(
        content=fileobj.getvalue().encode('utf-8'),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=SpiderFoot-search.csv",
            "Pragma": "no-cache",
        },
    )
